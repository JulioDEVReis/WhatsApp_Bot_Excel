"""
1. Pegar a planilha do cliente com dados de clientes dele contendo nome, telefone, data ultima compra, quantidade de
compras efetuadas, data aniversário
OBS: Numeros de telefone precisam ter codigo do país e ddd
OBS: Usuario precisará logar no WhatsApp Web cada vez que rodar o aplicativo. https://web.whatsapp.com/
2. Enviar mensagem automática no WhatsApp Web parabenizando pelo aniversário e oferecendo um desconto especial de
10% em qualquer compra naquela data
3. Enviar mensagem 15 dias após a compra pedindo feedback do cliente
4. Enviar mensagem 6 meses após a ultima compra oferecendo serviço de fotografia
5. Enviar mensagem após 3 compras agradecendo pela preferencia e dando desconto de 10% na próxima compra, e após
6 compras presenteando com um lindo album

Tecnologias usadas:
Automatização do Navegador - selenium.webdriver (pip install)
Adequação das datas - datetime (pip install)
Inclusão de pausas - time.sleep
Automatização da leitura de dados de uma planilha - Openpyxl (pip install)

Link WhatsApp Web: https://web.whatsapp.com/send?phone=&text= (texto codificado)
"""
# Importação da biblioteca para manipulação de dados em uma planilha excel
from openpyxl import load_workbook
# Biblioteca para datas
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
# Biblioteca de formatação de texto para links
from urllib.parse import quote
# Biblioteca para interação com a pagina web
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
# Biblioteca para as pausas entre tarefas
from time import sleep

# O programa aguardará 3 minutos até iniciar para dar tempo do computador inicializar
# sleep(180)

# Atrbuição do navegador escolhido
navegador = webdriver.Edge()
# abertura da pagina WhatsApp Web
navegador.get('https://web.whatsapp.com/')
# verificação de estamos logados no WhatsApp:
while len(navegador.find_elements(By.ID, "side")) == 0:
    sleep(1)


# Criação da função para enviar mensagem aos clientes via WhatsApp Web
def enviar_mensagem(mensagem):
    link = f"https://web.whatsapp.com/send?phone=55{telefone}&text={quote(mensagem)}"
    # Abertura do navegador com o link(parâmetro na função) - Selenium
    navegador.get(link)
    # Pausa de 10 segundos para iniciar a função, e dar tempo de abrir o navegador no computador
    sleep(15)
    # Verificar se está logado no WhatsApp Web
    while len(navegador.find_elements(By.ID, "side")) == 0:
        sleep(1)
    # enviar efetivamente a mensagem - pyautogui
    navegador.find_element(by=By.XPATH,value='/html/body/div[1]/div/div[2]/div[4]/div/footer/div[1]/div/span[2]/div/div[2]/div[2]/button').click()
    # Pausa de 15 segundos para proxima mensagem. Assim evita SPAM e bloqueio
    sleep(15)

# Formatar a data de hoje, de datetime.today() para o padrão que está na planilha
data_hoje = datetime.today().strftime('%d/%m/%Y')
data_hoje = datetime.strptime(data_hoje, '%d/%m/%Y')
# Acesso a planilha do cliente com openpyxl
planilha_cliente = load_workbook('clientes_estudio.xlsx')
# Acesso à aba que queremos extrair os dados na planilha Excel
aba_clientes = planilha_cliente['Clientes']

# Iteração sobre cada cliente dentro da aba, iniciando pela linha 5, que é o primeiro cliente na planilha
for cliente in aba_clientes.iter_rows(min_row=5):
    nome = cliente[0].value
    telefone = cliente[1].value
    data_ultima_compra = cliente[2].value
    quantidade_compras = cliente[3].value
    data_aniversario = cliente[4].value

    # uso de try pois os telefones podem não ser válidos ou não estarem cadastrados no WhatsApp
    try:
        # condição - se a data do aniversário for hoje, enviar mensagem de aniversário
        if data_aniversario == data_hoje:
            mensagem_aniversario = f"Olá {nome}, td bem? Gostaria de te parabenizar pela data de hoje!! Tudo de bom para Você! Felicidades e Saúde sempre! Somente hoje você poderá comprar qualquer item da nossa loja ou serviço de fotografia com 10% de desconto!"
            # função para envio da mensagem
            enviar_mensagem(mensagem_aniversario)

        # condição - se a data de hoje for igual 15 dias após a data da ultima compra, enviar mensagem de remarketing
        if data_hoje == data_ultima_compra + timedelta(days=15):
            mensagem_fidelizacao = f"Olá {nome}, como tem passado? Hoje tem 30 dias que entregamos nosso trabalho para você? O que você achou dele? Teria algo a acrescentar ou você gostou completamente do resultado?"
            # função para envio da mensagem
            enviar_mensagem(mensagem_fidelizacao)

        # condição - se a data de hoje for igual 6 meses após a data da ultima compra, enviar mensagem de remarketing
        if data_hoje == data_ultima_compra + relativedelta(months=6):
            mensagem_remarketing = f"Olá {nome}, quanto tempo... como vocês estão? Espero que tudo bem com vcs! Ahhh, boas lembranças são eternizadas pela fotografia, como vc já sabe. Que tal fazermos mais uma sessão fotográfica? Ou que tal aquela caneca, almofada ou calendário com as fotos que já fizemos para sempre lembrar daquele momento?"
            # função para envio da mensagem
            enviar_mensagem(mensagem_remarketing)

        # condição - se a quantidade de compras da cliente for igual a maior a 3 e menor que 6, enviar mensagem de cliente fiel básico
        if quantidade_compras >= 3 or quantidade_compras < 6:
            mensagem_cliente_fiel_basico = f"Olaaaa {nome}. Quero agradecer imensamente pelas suas compras aqui comigo. Fico feliz em saber que faço parte, indiretamente, da eternização de seus momentos mais especiais. Em forma de agradecimento, gostaria de presentear vocês com 15% de desconto na próxima compra, em qualquer produto."
            enviar_mensagem(mensagem_cliente_fiel_basico)
            # Se a data de hoje for igual 3 meses após a data da ultima compra, sendo cliente fiel básico, enviar mensagem de remarketing
            if data_hoje == data_ultima_compra + relativedelta(months=3):
                mensagem_remarketing_cliente_fiel = f"Oie!! Como vcs estão, {nome}? Lembrando que você tem conosco um desconto de 15% por ter sido sempre um cliente fiel a nós... sempre comprando conosco. Que tal usar seu desconto agora?"
                enviar_mensagem(mensagem_remarketing_cliente_fiel)

        # condição - se a quantidade de compras da cliente for igual a maior a 6, enviar mensagem de cliente fiel ouro
        if quantidade_compras >= 6:
            mensagem_cliente_fiel_ouro = f"Olaaaa {nome}. Nossa... já foram tantas vezes que você comprou algo aqui comigo que perdi as contas kkkkkkk. Estou muito feliz com todos esses momentos compartilhados. Quero agradecer mais uma vez por isso e então presentear vocês com um lindo album 15x15cm. Me chama que eu passo os detalhes! Parabéns pelo presente!"
            enviar_mensagem(mensagem_cliente_fiel_ouro)

    except Exception as e:
        print(f"não foi possivel enviar mensagem para {nome}: {e}")
        # abrir arquivo erros.css, 'a' de append para acrescentar dados, newline='' para não termos erros de formatação, encoding para formatar o texto
        with open('erros.csv','a',newline='',encoding='utf-8') as arquivo:
            arquivo.write(f"{nome},{telefone}")

navegador.quit()
